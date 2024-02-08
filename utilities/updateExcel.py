import os

import openpyxl

from utilities.outputVariables import *


class updateExcel:
    # toMail = ReadTestData.toMail()
    # Hfilename=[]
    # efile=[]
    @staticmethod
    def writedata(path, new_path, test, sheet, rownumStart, rownumFinish, columnno, data):
        wb = openpyxl.load_workbook(path)  # Workbook
        ws = wb[sheet]  # selecting the active sheet

        count = 0  # Variable to point to the position of list

        for i in range(rownumStart, rownumFinish + 1):
            temp = ws.cell(row=i, column=columnno)  # Loop through the required rows
            temp.value = data[count]  # inserting value in cell
            count += 1

        # sheet_name = str(datetime.now().strftime("%Y%m%d%H%M"))  # Current date time in local system
        file = new_path + test + str(str_localtime) + '.xlsx'

        wb.save(file)  # saving the changes

    @staticmethod
    def writedata_screenshot(path, new_path, test, sheet, rownumStart, rownumFinish, columnno, data, screenshot):
        wb = openpyxl.load_workbook(path)  # Workbook
        ws = wb[sheet]  # selecting the active sheet

        count_1 = 0  # Variable to point to the position of list
        count_2 = 0

        for i in range(rownumStart, rownumFinish + 1):
            temp = ws.cell(row=i, column=columnno)  # Loop through the required rows
            temp.value = data[count_1]  # inserting value in cell
            count_1 += 1

        for i in range(rownumStart, rownumFinish + 1):
            temp = ws.cell(row=i, column=columnno + 1)  # Loop through the required rows
            temp.value = screenshot[count_2]  # inserting value in cell
            count_2 += 1

        sheet_name = str(datetime.now().strftime("%Y%m%d%H%M"))  # Current date time in local system
        file = new_path + test + sheet_name + '.xlsx'

        wb.save(file)  # saving the changes

    @staticmethod
    def writedata_all(path, new_path, test, sheet, l1, l2, device_info):
        if os.path.exists(new_path):
            print("Excel Folder is available")
        else:
            os.mkdir("Excel")
            print("Excel Folder Created")
        wb = openpyxl.load_workbook(path)  # Workbook
        ws = wb[sheet]  # selecting the active sheet
        print(l1)
        print(l2)

        # Variable to point to the position of list
        for z in range(len(l1)):
            count = 0
            x = l1[z]
            y = l2[z]
            print(x, y)
            i, j, k = y[0], y[1], y[2]
            if len(x) == j - i + 1:
                for s in range(i, j + 1):
                    temp = ws.cell(row=s, column=k)  # Loop through the required rows
                    temp.value = x[count]  # inserting value in cell
                    count += 1
            else:
                print('Out of range for test ' + str(i) + ' to ' + str(j))

        wd = wb['Coverage Summary']
        (wd.cell(row=3, column=10)).value = device_info[0]
        (wd.cell(row=4, column=10)).value = device_info[1]
        (wd.cell(row=4, column=3)).value = device_info[2]
        (wd.cell(row=5, column=3)).value = device_info[3]

        efile = new_path + test + str(str_localtime) + '.xlsx'

        wb.save(efile)  # saving the changes
        '''
        print("Excel File Name as below")
        print(efile)
        print("HTML file Name as below")
        Hfilename1 = 'Reports/' 'JIO_HGWAutomation_HTMLReport_' + str(str_localtime) + '.html'
        Hfilename.append(Hfilename1)
        print(str(Hfilename))
        commonmethod.mailAlert(str(toMail), str(Hfilename))
        commonmethod.mailAlert(str(toMail), efile)'''

    # test_name = JIO_HGWAutomation_ExcelReport_

    @staticmethod
    def writedata_coverageSummary(file, device_info):
        print(device_info)
        print(file)

        wb = openpyxl.load_workbook(file)  # Workbook
        wd = wb['Coverage Summary']
        (wd.cell(row=3, column=10)).value = device_info[0]
        (wd.cell(row=4, column=10)).value = device_info[1]
        (wd.cell(row=4, column=3)).value = device_info[2]
        (wd.cell(row=5, column=3)).value = device_info[3]
        wb.save(file)
